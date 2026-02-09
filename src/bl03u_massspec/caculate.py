import os
import re
import sys

import openpyxl
import matplotlib.pyplot as plt
from numpy import trapz
import pandas as pd
import time
from PyQt6.QtWidgets import QFileDialog

NUMBER_REGEX = re.compile(r'[-+]?(?:\d+\.\d*|\d+|\.\d+)(?:[eE][-+]?\d+)?')


# 获取excel中某表、某行、某列单元格值的函数
# 输入表格名、列数、行数，索取该单元格的值
def get_value(sheet_name, col_number, row_number):
    return sheet_name.cell(column=col_number, row=row_number).value


# 获取卡峰数据的函数
# 从卡峰配置文件中读取已经设定好的物种的m/z以及峰的peak,start,end值
def get_configuration():
    species_info = {}
    configuration_path = os.getcwd()

    try:
        wb = openpyxl.load_workbook(configuration_path + '//data' + '//卡峰配置文件.xlsx', data_only=True)
        ws = wb['卡峰配置']
    except FileNotFoundError:
        print("未找到卡峰配置文件！")

    for row in range(2, ws.max_row + 1):
        species_name = get_value(ws, 1, row)  # 2, 物种名称（需要修改保存图文件路径的代码） 1, 质量数   注意数据格式！！！
        if species_name is None:
            continue
        species_info.setdefault(species_name, {})
        for col in range(3, 6):
            position_name = get_value(ws, col, 1)
            position_value = get_value(ws, col, row)
            species_info[species_name].setdefault(position_name, position_value)

    return species_info


# 获取定标三参数的函数
# 并且向参数历史中写入参数调用历史
def get_para():
    para = []
    configuration_path = os.getcwd()
    try:
        para_path = os.path.join(configuration_path, 'data', '定标参数.txt')
        with open(para_path) as f:
            lines = f.readlines()
            para = [float(x.strip()) for x in lines]
    except FileNotFoundError:
        print("未找到定标参数文件！")
    # 将列表转换为字符串
    para_str = ', '.join(map(str, para))
    # 写入参数历史文件
    try:
        with open(os.path.join(configuration_path, 'data', '参数历史.txt'), 'a') as f:
            now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            f.write(now + '\n')  # 写入当前时间
            f.write(para_str)  # 写入转换后的参数字符串
    except FileNotFoundError:
        print("未找到参数历史文件！")
    return para


# 读取文件夹中所有txt文件的函数
# 输入文件夹路径，输出文件夹中所有txt文件的路径

def get_file_path(folder_path):
    for folder_name, sub_folders, filenames in os.walk(folder_path):
        for i in range(len(filenames)):
            filenames[i] = os.path.join(folder_path, filenames[i])

        return filenames


# folder_path = './data/raw/C6F11O2H/11.0eV'
# file = get_file_path(folder_path)
# print(file)


# 获取txt中信号信息
# 输入数据储存路径以及是否选择头文件矫正模式：如果选择矫正模式，则输出对应矫正后的信号，否则输出信号的原始数据。


def get_signal(file_path, choice):
    with open(file_path) as f:
        intensity = f.readlines()
        intensity = [x.strip() for x in intensity]

        # 无头文件模式
        intensity_info = intensity
        head_info = intensity

        # 头文件矫正模式,如果是催化质谱，有10行头文件，如果是TOF1,有7行头文件。
        intensity_info = intensity[10:]
        head_info = intensity[:10]
    if choice == 'intensity':
        return intensity_info
    elif choice == 'head':
        return head_info
    else:
        return '输入了错误的请求信息'


# intens = get_signal('./data/raw/C6F11O2H/11.0eV/C23072501-0000.txt', 'head')
# print(intens)


# 识别头文件中的数据
def get_head_info(head_list):
    head_list_ab = []
    for head in head_list:
        mo = NUMBER_REGEX.search(head)
        head_info = mo.group() if mo else "1"

        # # 不矫正模式
        # head_list_ab.append(1)

        # 头文件流强矫正模式
        head_list_ab.append(head_info)

    return head_list_ab


# head_list = get_head_info(get_signal('./data/raw/C6F11O2H/11.0eV/C23072501-0000.txt', 'head'))
# print(head_list)


# 定义积分函数
# 输入包含物种的m/z以及峰的peak,start,end值的配置信息输出积分后的峰面积
def integrate(info, list_intensity, mode='integrate'):
    list_peak_area = {}
    list_peak_height = {}
    for species in info:
        start = info[species]['Start']
        end = info[species]['End']
        list_tof = [x for x in range(start, end + 1)]
        list_intensity_slice = list_intensity[start: end + 1]
        list_intensity_slice = list(map(int, list_intensity_slice))
        integrate_peak = trapz(list_intensity_slice, list_tof, dx=0.01)
        peak_height = max(list_intensity_slice)
        list_peak_height.setdefault(species, peak_height)
        list_peak_area.setdefault(species, integrate_peak)
    if mode == 'integrate':
        return list_peak_area
    else:
        return list_peak_height


# info = get_configuration()
# # print(info)
# list_intensity = get_signal('./data/raw/C6F11O2H/12.0eV/C23072509-0000.txt', 'intensity')
# # print(list_intensity)
# area = integrate(info, list_intensity)
# print(area)


def calculate(path):
    # path = './data/raw/C6F11O2H/12.0eV'  # 文件夹路径, 需要修改
    file_paths = get_file_path(path)
    list_integrate = []
    for file_path in file_paths:
        intensity = get_signal(file_path, 'intensity')
        head_list = get_head_info(get_signal(file_path, 'head'))
        integrate_result = integrate(get_configuration(), intensity)
        # 用流强矫正，如用TOF-1，需要注释掉
        for result in integrate_result:
            integrate_result[result] = integrate_result[result] / float(head_list[1])
        list_integrate.append(integrate_result)
    return list_integrate


# print(calculate('./data/raw/C6F11O2H/12.0eV'))


# 出数据的函数
def get_result(path):
    list_integrate = calculate(path)
    list_number = []
    dic_area = {}

    for i, integrate_result in enumerate(list_integrate):
        list_number.append(i)
        for species in integrate_result:
            dic_area.setdefault(species, []).append(integrate_result[species])
    return dic_area
    # 原本的错误代码
    # for i, integrate_result in enumerate(list_integrate):
    #     list_number.append(i)
    #     for species in integrate_result:
    #         dic_area.setdefault(species, []).append(integrate_result[species])
    # return dic_area


# print(get_result())


# 生成图片的函数
def generate_graph(x, y, label, title, xlabel, ylabel, path_save):
    fig = plt.figure()
    plt.plot(x, y, label=label)
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.legend()
    fig.savefig(path_save)
    plt.close(fig)


# 获取文件头信息 0-能量 1-光电流 2-束流强度 3-位置 4-采谱时间 5- 6-温度 7，8，9-腔体真空
def get_yaxis(path, index):
    # path = './data/raw/C6F11O2H/12.0eV'  # 文件夹路径, 需要修改
    file_paths = get_file_path(path)
    head_list = []
    for file_path in file_paths:
        head_result = get_head_info(get_signal(file_path, 'head'))
        head_list.append(head_result[index])

    return head_list


# print(yaxis())

# 出图函数
# 输入单谱时间，输出信号随时间变化图
# def get_graph(path, dic):
#     for species in dic:
#         print(str(species) + '趋势正在出图中')
#         list_area = dic[species]
#
#         list_temperature = get_yaxis(path, 6)
#         generate_graph(
#             x=list_temperature, y=list_area, label=' ', title=str(species),
#             xlabel='Temperature(℃)', ylabel='area', path_save=os.getcwd() + '\\image' + '\\Temperature' + '\\'
#                                                               + str(species) + '.png'
#             # 保存路径，保存到当前文件夹下的image文件夹中
#         )


# dic = get_result()
# get_graph(dic)
# for species in dic:
#     print(dic[species])


# 主程序
def main():
    # 记录初始时间
    start_time = time.time()
    default_path = os.path.join(os.getcwd(), 'data', 'raw', 'C6F11O2H', '11.5eV')
    path = sys.argv[1] if len(sys.argv) > 1 else default_path
    if not os.path.isdir(path):
        print(f"路径不存在: {path}")
        print("用法: python caculate.py <数据目录>")
        return

    # 打印采谱时间等配置信息
    dic = get_result(path)
    first_key = list(dic.keys())[0]
    first_value = dic[first_key]
    len_data = len(first_value)
    print('数据总量' + str(len_data))

    # get_graph(path, dic)
    for species in dic:
        print(str(species) + '趋势正在出图中')
        list_area = dic[species]
        output_dir = os.path.join(os.getcwd(), 'image', 'Temperature')
        os.makedirs(output_dir, exist_ok=True)

        list_temperature = get_yaxis(path, 6)
        generate_graph(
            x=list_temperature, y=list_area, label=' ', title=str(species),
            xlabel='Temperature(℃)', ylabel='area', path_save=os.path.join(output_dir, f'{species}.png')
            # 保存路径，保存到当前文件夹下的image文件夹中
        )

    df = pd.DataFrame(dic)
    file_path, _ = QFileDialog.getSaveFileName(
        None,
        "保存数据",
        "",
        "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*.*)"
    )
    if file_path:
        if file_path.endswith('.xlsx'):
            df.to_excel(file_path, index=False)
        elif file_path.endswith('.csv'):
            df.to_csv(file_path, index=False)
        else:
            df.to_excel(file_path, index=False)
        print(f"数据已保存至：{file_path}")

    # 记录结束时间，计算运行时间总长，并打印
    end_time = time.time()
    time_difference = end_time - start_time
    print(f"本次运行总时间: {time_difference} 秒")


if __name__ == '__main__':
    main()
